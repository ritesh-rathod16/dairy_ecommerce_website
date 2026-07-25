"""
Renders the analytics data dict (from app.services.analytics.compute_analytics)
as a PDF or Excel file. Both take the exact same computed dict the JSON
endpoint returns, so an export can never show different numbers than the
dashboard did.
"""
import io

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _rs(amount) -> str:
    return f"Rs. {amount:,.2f}"


def build_analytics_pdf(data: dict, merchant_name: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm, leftMargin=18 * mm, rightMargin=18 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=18)
    small = ParagraphStyle("Small2", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    h2 = ParagraphStyle("H2", parent=styles["Heading3"], spaceBefore=12, spaceAfter=4)

    story = [
        Paragraph(f"{merchant_name} — Analytics Report", title_style),
        Paragraph(f"{data['range']['from']} to {data['range']['to']}", small),
        Spacer(1, 12),
    ]

    money_keys = {"revenue", "average_order_value", "cod_pending_amount"}
    kpi_rows = [["Metric", "Value", "vs previous period"]]
    for key, label in [
        ("revenue", "Revenue"), ("orders", "Orders"), ("average_order_value", "Avg Order Value"),
        ("unique_customers", "Unique Customers"), ("repeat_customers", "Repeat Customers"),
        ("cancelled_orders", "Cancelled Orders"), ("cod_pending_amount", "COD Pending"),
    ]:
        k = data["kpis"][key]
        val = _rs(k["value"]) if key in money_keys else k["value"]
        change = f"{k['change_pct']:+.1f}%" if k.get("change_pct") is not None else "—"
        kpi_rows.append([label, str(val), change])

    kpi_table = Table(kpi_rows, colWidths=[200, 150, 120])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4332")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBF6EC")]),
    ]))
    story.append(kpi_table)

    story.append(Paragraph("Category sales", h2))
    if data["category_sales"]:
        rows = [["Category", "Qty sold", "Revenue", "% of total"]]
        for c in data["category_sales"]:
            rows.append([c["name"], str(c["qty"]), _rs(c["revenue"]), f"{c['pct']}%"])
        t = Table(rows, colWidths=[180, 90, 110, 90])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4332")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9), ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E5")),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No sales in this range.", small))

    story.append(Paragraph("Top customers", h2))
    if data["top_customers"]:
        rows = [["Customer", "Orders", "Total spent"]]
        for c in data["top_customers"][:10]:
            rows.append([c["name"], str(c["order_count"]), _rs(c["total_spent"])])
        t = Table(rows, colWidths=[220, 100, 150])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4332")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9), ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E5")),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No customers in this range.", small))

    story.append(Paragraph("Order status summary", h2))
    status_rows = [["Status", "Count"]] + [[k.replace("_", " ").title(), str(v)] for k, v in data["order_status_summary"].items()]
    t = Table(status_rows, colWidths=[200, 100])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4332")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9), ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5E5")),
    ]))
    story.append(t)

    doc.build(story)
    return buf.getvalue()


def build_analytics_excel(data: dict) -> bytes:
    wb = Workbook()
    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def _header(ws, row, cols):
        for i, c in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=i, value=c)
            cell.fill = header_fill
            cell.font = header_font

    ws = wb.active
    ws.title = "KPIs"
    _header(ws, 1, ["Metric", "Value", "Change vs previous period (%)"])
    row = 2
    for key, label in [
        ("revenue", "Revenue"), ("orders", "Orders"), ("average_order_value", "Avg Order Value"),
        ("unique_customers", "Unique Customers"), ("repeat_customers", "Repeat Customers"),
        ("active_deliveries", "Active Deliveries"), ("cancelled_orders", "Cancelled Orders"),
        ("cod_pending_amount", "COD Pending Amount"), ("cod_collected_today", "COD Collected Today"),
    ]:
        k = data["kpis"][key]
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=k["value"])
        ws.cell(row=row, column=3, value=k["change_pct"] if k.get("change_pct") is not None else "N/A")
        row += 1

    ws2 = wb.create_sheet("Revenue by day")
    _header(ws2, 1, ["Date", "Revenue", "Orders"])
    for i, d in enumerate(data["revenue_orders_series"], start=2):
        ws2.cell(row=i, column=1, value=d["date"])
        ws2.cell(row=i, column=2, value=d["revenue"])
        ws2.cell(row=i, column=3, value=d["orders"])

    ws3 = wb.create_sheet("Category sales")
    _header(ws3, 1, ["Category", "Qty sold", "Revenue", "% of total"])
    for i, c in enumerate(data["category_sales"], start=2):
        ws3.cell(row=i, column=1, value=c["name"])
        ws3.cell(row=i, column=2, value=c["qty"])
        ws3.cell(row=i, column=3, value=c["revenue"])
        ws3.cell(row=i, column=4, value=c["pct"])

    ws4 = wb.create_sheet("Top customers")
    _header(ws4, 1, ["Customer", "Orders", "Total spent", "Last order"])
    for i, c in enumerate(data["top_customers"], start=2):
        ws4.cell(row=i, column=1, value=c["name"])
        ws4.cell(row=i, column=2, value=c["order_count"])
        ws4.cell(row=i, column=3, value=c["total_spent"])
        ws4.cell(row=i, column=4, value=str(c["last_order_at"]))

    ws5 = wb.create_sheet("Order status")
    _header(ws5, 1, ["Status", "Count"])
    for i, (k, v) in enumerate(data["order_status_summary"].items(), start=2):
        ws5.cell(row=i, column=1, value=k)
        ws5.cell(row=i, column=2, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
